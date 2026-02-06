"""Excel report generation."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config.settings import get_settings

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export model results to Excel workbook.

    Sheets:
    1. Predictions - All games with model spread vs Vegas opener
    2. Value Plays - Filtered to significant edge games
    3. Power Ratings - All teams ranked by overall rating
    4. Components - Detailed breakdown per game
    """

    def __init__(self, output_dir: Optional[Path] = None):
        """Initialize Excel exporter.

        Args:
            output_dir: Output directory for Excel files
        """
        settings = get_settings()
        self.output_dir = output_dir or settings.outputs_dir

    def _style_header(self, ws, num_cols: int) -> None:
        """Apply header styling to worksheet."""
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def _style_value_plays(self, ws, df: pd.DataFrame) -> None:
        """Apply conditional styling to value plays sheet."""
        # Highlight high-edge plays
        high_edge_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )

        if "edge" in df.columns:
            edge_col = df.columns.get_loc("edge") + 1

            for row_idx, value in enumerate(df["edge"], start=2):
                if value is not None and abs(value) >= 5:
                    for col in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=col).fill = high_edge_fill

    def _auto_column_width(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_borders(self, ws) -> None:
        """Add borders to all cells with data."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

    def create_predictions_sheet(
        self,
        wb: Workbook,
        predictions_df: pd.DataFrame,
    ) -> None:
        """Create predictions sheet.

        Args:
            wb: Workbook to add sheet to
            predictions_df: DataFrame with predictions
        """
        ws = wb.create_sheet("Predictions")

        # Select and order columns
        columns = [
            "home_team",
            "away_team",
            "spread",
            "favorite",
            "vegas_spread",
            "edge",
            "home_win_prob",
            "confidence",
        ]

        available_cols = [c for c in columns if c in predictions_df.columns]
        df = predictions_df[available_cols].copy()

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx + 1, column=c_idx)
                if isinstance(value, float):
                    cell.value = round(value, 1)
                    cell.number_format = "0.0"
                else:
                    cell.value = value

        self._style_header(ws, len(available_cols))
        self._auto_column_width(ws)
        self._add_borders(ws)

    def create_value_plays_sheet(
        self,
        wb: Workbook,
        value_plays_df: pd.DataFrame,
    ) -> None:
        """Create value plays sheet.

        Args:
            wb: Workbook to add sheet to
            value_plays_df: DataFrame with value plays
        """
        ws = wb.create_sheet("Value Plays")

        if value_plays_df.empty:
            ws.cell(row=1, column=1, value="No value plays identified this week")
            return

        columns = [
            "team",
            "side",
            "model_spread",
            "vegas_spread",
            "edge",
            "confidence",
            "analysis",
        ]

        available_cols = [c for c in columns if c in value_plays_df.columns]
        df = value_plays_df[available_cols].copy()

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx + 1, column=c_idx)
                if isinstance(value, float):
                    cell.value = round(value, 1)
                    cell.number_format = "0.0"
                else:
                    cell.value = value

        self._style_header(ws, len(available_cols))
        self._style_value_plays(ws, df)
        self._auto_column_width(ws)
        self._add_borders(ws)

    def create_power_ratings_sheet(
        self,
        wb: Workbook,
        ratings_df: pd.DataFrame,
    ) -> None:
        """Create power ratings sheet.

        Follows the JP+ Power Ratings Display Protocol (see CLAUDE.md):
        Rank | Team | Overall | Offense (rank) | Defense (rank) | Special Teams (rank)

        Args:
            wb: Workbook to add sheet to
            ratings_df: DataFrame with team ratings
        """
        ws = wb.create_sheet("Power Ratings")

        # Add rank column
        df = ratings_df.copy()
        df.insert(0, "rank", range(1, len(df) + 1))

        columns = ["rank", "team", "overall", "offense", "defense"]
        available_cols = [c for c in columns if c in df.columns]
        df = df[available_cols]

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx + 1, column=c_idx)
                if isinstance(value, float):
                    cell.value = round(value, 2)
                    cell.number_format = "0.00"
                else:
                    cell.value = value

        self._style_header(ws, len(available_cols))
        self._auto_column_width(ws)
        self._add_borders(ws)

        # Highlight top 25
        top_fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        for row in range(2, min(27, len(df) + 2)):
            for col in range(1, len(available_cols) + 1):
                ws.cell(row=row, column=col).fill = top_fill

    def create_components_sheet(
        self,
        wb: Workbook,
        predictions_df: pd.DataFrame,
    ) -> None:
        """Create components breakdown sheet.

        Args:
            wb: Workbook to add sheet to
            predictions_df: DataFrame with full component breakdown
        """
        ws = wb.create_sheet("Components")

        columns = [
            "home_team",
            "away_team",
            "spread",
            "base_margin",
            "hfa",
            "situational",
            "travel",
            "altitude",
            "special_teams",
            "finishing_drives",
            "luck_adj",
        ]

        available_cols = [c for c in columns if c in predictions_df.columns]
        df = predictions_df[available_cols].copy()

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx + 1, column=c_idx)
                if isinstance(value, float):
                    cell.value = round(value, 2)
                    cell.number_format = "0.00"
                else:
                    cell.value = value

        self._style_header(ws, len(available_cols))
        self._auto_column_width(ws)
        self._add_borders(ws)

    def export(
        self,
        predictions_df: pd.DataFrame,
        value_plays_df: pd.DataFrame,
        ratings_df: pd.DataFrame,
        year: int,
        week: int,
        filename: Optional[str] = None,
    ) -> Path:
        """Export all data to Excel workbook.

        Args:
            predictions_df: Full predictions with comparison
            value_plays_df: Filtered value plays
            ratings_df: Team power ratings
            year: Season year
            week: Week number
            filename: Custom filename (optional)

        Returns:
            Path to created file
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d")
            filename = f"cfb_predictions_{year}_week{week}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create sheets
        self.create_predictions_sheet(wb, predictions_df)
        self.create_value_plays_sheet(wb, value_plays_df)
        self.create_power_ratings_sheet(wb, ratings_df)
        self.create_components_sheet(wb, predictions_df)

        # Add summary sheet as first sheet
        summary = wb.create_sheet("Summary", 0)
        summary.cell(row=1, column=1, value="CFB Power Ratings Model")
        summary.cell(row=1, column=1).font = Font(size=16, bold=True)

        summary.cell(row=3, column=1, value=f"Season: {year}")
        summary.cell(row=4, column=1, value=f"Week: {week}")
        summary.cell(
            row=5,
            column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )
        summary.cell(row=7, column=1, value=f"Total Games: {len(predictions_df)}")
        summary.cell(row=8, column=1, value=f"Value Plays: {len(value_plays_df)}")
        summary.cell(row=9, column=1, value=f"Teams Rated: {len(ratings_df)}")

        self._auto_column_width(summary)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported Excel report to {filepath}")

        return filepath
